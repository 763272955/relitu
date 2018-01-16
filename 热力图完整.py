# -*- coding:utf-8 -*-

import urllib2
import  time
import datetime
import csv
import sys
import openpyxl
# reload(sys)
# sys.setdefaultencoding('utf8')

nowtime = datetime.datetime.now().strftime('%Y%m%d')
nowtimel = nowtime+".csv"
nowtime_ = nowtime+".xlsx"
nowtime = nowtime+".txt"
dic_IP={}
list_num=[]
def Csv2Xlxs(nowtimel, nowtime_):
    wb = openpyxl.Workbook()
    ws = wb.active
    f = open('inputfile/'+nowtimel)
    reader = csv.reader(f, delimiter=',')
    for row in reader:
        row_container = []
        count = 1
        for cell in row:
            try:
                cell = int(float(cell))
                row_container.append(cell)
                continue
            except:
                pass
            row_container.append(cell.decode('gbk').encode('utf-8'))
        ws.append(row_container)
    f.close()
    wb.save('inputfile/'+nowtime_)
def search_database():
    wb = openpyxl.load_workbook('inputfile/'+nowtime_)
    ws = wb.get_sheet_by_name('Sheet')
    for row in list(ws.rows)[1:]:
        dic_IP[row[0].value]=row[1].value
    for key in dic_IP.keys():
        list_num.append(key)
    return list_num,dic_IP
def search_IP(list_num,dic_IP):
    # print list_num
    flag = 0
    file= open('outputfile/'+nowtime,'w')

    for IP in list_num:
        number = dic_IP[IP]
        number=str(number)
        IP = unicode(IP).encode('utf8')
        res =urllib2.urlopen("http://api.map.baidu.com/geocoder?address=%s&output=json&key=6eea93095ae93db2c77be9ac910ff311&city=%s"  % (IP,IP))
        a = res.read()
        # print a
        zidian = eval(a)
        # print zidian
        flag = flag + 1
        if   len(zidian['result'])==0:
            print '%s is not in area' % IP
            continue
        else:
            a = str(zidian['result']['location']['lat'])
            b = str(zidian['result']['location']['lng'])
            str_temp = '{"lat":' + a + ',"lng":' + b + ',"count":' + number + '}, \n'
            print a, b, type(a), type(b)
        file.write(str_temp)
    file.close()
def hebing_():
    lis = [u"国家地区", u"事件数"]
    nowtime = datetime.datetime.now().strftime('%Y%m%d')
    nowtime_ = nowtime + ".xlsx"

    wb = openpyxl.load_workbook('inputfile/' + nowtime_)
    wk_ = wb.create_sheet('sheet2')
    ws = wb.get_sheet_by_name('Sheet')
    dic = {}
    for row in list(ws.rows)[1:]:
        # dic[row[0].value] = int(row[1].value)
        if row[0].value not in dic.keys():
            dic[row[0].value] = int(row[1].value)
        else:
            dic[row[0].value] += int(row[1].value)
    wc = wb.get_sheet_by_name('sheet2')
    wc.append(lis)
    for key in dic.keys():
        wc.append((key, dic[key]))
    wb.save('inputfile/' + nowtime_)
# Csv2Xlxs(nowtimel,nowtime_)
# hebing_()
list_num, dic_IP = search_database()
search_IP(list_num,dic_IP)


#http://developer.baidu.com/map/jsdemo.htm#c1_15